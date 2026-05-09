# dashboard.py
# ──────────────────────────────────────────────────────────────
# Dashboard Streamlit — Operación UBER 2025 / 2026
# Usa business_rules.py para cargar y transformar datos.
# ──────────────────────────────────────────────────────────────
import os
import re
import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go

from business_rules import (
    load_ingresos,
    load_egresos,
    add_yearweek,
    yearweek_label,
    get_analisis_inversiones,
    get_gastos,
    get_inversiones_egreso,
    conductores_por_semana,
    auditoria_signos,
    auditoria_egresos,
    PAGO_SEMANAL_DEFAULT,
    CONCEPTOS_INVERSION,
)

# ═══════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Dashboard Uber 2025-2026",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═══════════════════════════════════════════════════════════════
# CUSTOM THEME / CSS
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<style>
    /* KPI cards */
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 1px solid #0f3460;
        border-radius: 12px;
        padding: 16px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.3);
    }
    div[data-testid="stMetric"] label {
        color: #a8b8d8 !important;
        font-size: 0.85rem !important;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #e8f0fe !important;
        font-size: 1.4rem !important;
        font-weight: 700 !important;
    }
    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 10px 20px;
        font-weight: 600;
    }
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0d1b2a 0%, #1b2838 100%);
    }
    /* Dataframes */
    .stDataFrame {
        border-radius: 8px;
        overflow: hidden;
    }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# PLOTLY THEME
# ═══════════════════════════════════════════════════════════════
PLOT_TEMPLATE = "plotly_dark"
COLOR_INGRESOS = "#00d4aa"
COLOR_EGRESOS = "#ff6b6b"
COLOR_NETO = "#4ecdc4"
COLOR_PALETTE = px.colors.qualitative.Set2


def styled_bar(fig, money_axis="y"):
    """money_axis: 'y', 'x', 'both', or None."""
    fmt = "$,.0f"
    ya = dict(tickformat=fmt) if money_axis in ("y", "both") else {}
    xa = dict(tickformat=fmt) if money_axis in ("x", "both") else {}
    fig.update_layout(
        template=PLOT_TEMPLATE,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#c8d6e5"),
        margin=dict(l=60, r=20, t=40, b=40),
        legend=dict(orientation="h", y=-0.15),
        yaxis=ya,
        xaxis=xa,
    )
    return fig


def styled_line(fig, money_axis="y"):
    fmt = "$,.0f"
    ya = dict(tickformat=fmt) if money_axis in ("y", "both") else {}
    xa = dict(tickformat=fmt) if money_axis in ("x", "both") else {}
    fig.update_layout(
        template=PLOT_TEMPLATE,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#c8d6e5"),
        margin=dict(l=60, r=20, t=40, b=40),
        yaxis=ya,
        xaxis=xa,
    )
    return fig


# ════════════════════════════════════════════════════════
# EXCEL EXPORT — ANÁLISIS DE INVERSIÓN
# ════════════════════════════════════════════════════════
def _generar_excel_amortizacion(
    resumen_inv: "pd.DataFrame",
    tablas_inv: dict,
    df_egr_para_inv: "pd.DataFrame",
    kpis: dict,
) -> bytes:
    """Genera el archivo Excel en memoria con 4 hojas de análisis de inversión.

    Hojas:
        Amortizacion       — tabla semanal por vehículo
        Resumen            — métricas agregadas por vehículo
        KPIs               — indicadores globales del módulo
        BaseComprasVehiculo— egresos originales de compra
    """
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # ── Formatos numéricos ────────────────────────────────────
    FMT_MONEY = '$#,##0.00'
    FMT_PCT   = '0.00%'
    FMT_NUM   = '#,##0'

    # Columnas con formato moneda por hoja
    MONEY_COLS: dict[str, set] = {
        "Amortizacion": {"inversion_inicial", "pago_recuperacion", "saldo_restante"},
        "Resumen": {
            "capital_inicial", "pago_semanal", "monto_recuperado",
            "ingresos_totales",
        },
        "BaseComprasVehiculo": {"monto_real"},
    }
    # Columnas con formato porcentaje por hoja
    PCT_COLS: dict[str, set] = {
        "Resumen": {"roi_pct"},
    }

    # ── Estilo de encabezado ───────────────────────────────────
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", fgColor="0F3460")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ROW_ALIGN = Alignment(vertical="center")

    # ── Construir DataFrame de amortización (todas las llaves concatenadas) ⁐
    frames_am = []
    for llave, tabla in tablas_inv.items():
        capital = 0.0
        if not resumen_inv.empty and "capital_inicial" in resumen_inv.columns:
            mask = resumen_inv["llave"].astype(str) == llave
            if mask.any():
                capital = float(resumen_inv.loc[mask, "capital_inicial"].iloc[0])
        t = tabla.copy()
        t.insert(0, "llave", llave)
        t.insert(1, "inversion_inicial", capital)
        frames_am.append(t)

    df_amort = (
        pd.concat(frames_am, ignore_index=True)
        if frames_am
        else pd.DataFrame(columns=["llave", "inversion_inicial", "semana_num",
                                    "pago_recuperacion", "saldo_restante"])
    )

    # ── Construir DataFrame de KPIs ──────────────────────────────
    kpi_filas = []
    for metrica, info in kpis.items():
        kpi_filas.append({
            "Métrica": metrica,
            "Valor":   info["valor"],
            "Tipo":    info["tipo"],   # 'moneda' | 'numero' | 'porcentaje' | 'texto'
        })
    df_kpis = pd.DataFrame(kpi_filas)

    # ── Construir DataFrame de base de compras ────────────────────
    if "concepto" in df_egr_para_inv.columns and not df_egr_para_inv.empty:
        df_base = df_egr_para_inv[
            df_egr_para_inv["concepto"].astype(str).str.strip().str.upper()
            == "PAGO DE SUBASTA"
        ].copy()
    else:
        df_base = pd.DataFrame()

    # ── Escribir a Excel ─────────────────────────────────────
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_amort.to_excel(writer, sheet_name="Amortizacion",        index=False)
        resumen_inv.to_excel(writer, sheet_name="Resumen",           index=False)
        df_kpis.to_excel(writer,     sheet_name="KPIs",              index=False)
        df_base.to_excel(writer,     sheet_name="BaseComprasVehiculo", index=False)

        wb = writer.book

        for sheet_name in ["Amortizacion", "Resumen", "KPIs", "BaseComprasVehiculo"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            # Congelar primera fila
            ws.freeze_panes = "A2"

            # Leer encabezados
            headers = [cell.value for cell in ws[1]]
            money_set = MONEY_COLS.get(sheet_name, set())
            pct_set   = PCT_COLS.get(sheet_name, set())

            # Calcular ancho máximo por columna
            col_widths: list[int] = []
            for col_idx, header in enumerate(headers, start=1):
                max_len = len(str(header)) if header else 8
                for row in ws.iter_rows(
                    min_row=2, max_row=ws.max_row,
                    min_col=col_idx, max_col=col_idx,
                ):
                    for cell in row:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                col_widths.append(min(max_len + 4, 45))

            # Aplicar estilos encabezado + ancho + formato numérico
            for col_idx, header in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = col_widths[col_idx - 1]

                # Estilo encabezado
                hdr_cell = ws.cell(row=1, column=col_idx)
                hdr_cell.font      = HDR_FONT
                hdr_cell.fill      = HDR_FILL
                hdr_cell.alignment = HDR_ALIGN

                # Formato numérico de celdas de datos
                for data_row in ws.iter_rows(
                    min_row=2, max_row=ws.max_row,
                    min_col=col_idx, max_col=col_idx,
                ):
                    for cell in data_row:
                        cell.alignment = ROW_ALIGN
                        if header in money_set and isinstance(cell.value, (int, float)):
                            cell.number_format = FMT_MONEY
                        elif header in pct_set and isinstance(cell.value, (int, float)):
                            # roi_pct ya viene en puntos porcentuales (ej. -98.12)
                            # lo dividimos por 100 para que Excel lo muestre como %
                            cell.value         = cell.value / 100.0
                            cell.number_format = FMT_PCT

            # Fila de altura uniforme
            ws.row_dimensions[1].height = 20

            # Para KPIs, aplicar formato numérico según columna Tipo
            if sheet_name == "KPIs" and "Valor" in headers and "Tipo" in headers:
                valor_col = headers.index("Valor") + 1
                tipo_col  = headers.index("Tipo")  + 1
                for r in range(2, ws.max_row + 1):
                    tipo_val  = ws.cell(row=r, column=tipo_col).value
                    valor_cell = ws.cell(row=r, column=valor_col)
                    if tipo_val == "moneda":
                        valor_cell.number_format = FMT_MONEY
                    elif tipo_val == "porcentaje" and isinstance(valor_cell.value, (int, float)):
                        valor_cell.value         = valor_cell.value / 100.0
                        valor_cell.number_format = FMT_PCT
                    elif tipo_val == "numero":
                        valor_cell.number_format = FMT_NUM

    buffer.seek(0)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════
# SIDEBAR: CARGA DE DATOS
# ═══════════════════════════════════════════════════════════════
# Detectar si existe el archivo local (solo disponible en desarrollo)
_DEFAULT_FILE = "prueba.xlsx"
_local_file_exists = os.path.isfile(_DEFAULT_FILE)

with st.sidebar:
    st.markdown("## 📥 Datos")
    uploaded = st.file_uploader("Sube tu Excel", type=["xlsx", "xls"])

    if _local_file_exists:
        # Entorno de desarrollo: ofrecer archivo local como alternativa
        use_default = st.checkbox(
            "Usar archivo local (prueba.xlsx)",
            value=(uploaded is None),
        )
    else:
        # Producción (Render): solo upload — no hay archivo local
        use_default = False
        if uploaded is None:
            st.info("📂 Sube tu archivo Excel para comenzar.")

path = uploaded if uploaded is not None else (_DEFAULT_FILE if use_default else None)

if path is None:
    st.warning("Sube un archivo Excel para visualizar los datos.")
    st.stop()


@st.cache_data(show_spinner="Cargando datos...")
def load_all(file_path):
    df_i = load_ingresos(file_path)
    df_e = load_egresos(file_path)
    return df_i, df_e


try:
    df_ing_raw, df_egr_raw = load_all(path)
except Exception as e:
    st.error(f"Error al cargar datos: {e}")
    st.stop()

if len(df_ing_raw) == 0 and len(df_egr_raw) == 0:
    st.warning("No se encontraron datos en el archivo.")
    st.stop()

# Agregar YEARWEEK
df_ing_raw = add_yearweek(df_ing_raw)
df_egr_raw = add_yearweek(df_egr_raw)

# ═══════════════════════════════════════════════════════════════
# SIDEBAR: FILTROS
# ═══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("---")
    st.markdown("## 🎛️ Filtros")

    # Año
    all_years = sorted(set(
        list(df_ing_raw["año"].dropna().unique()) +
        list(df_egr_raw["año"].dropna().unique())
    ))
    if len(all_years) == 0:
        st.error("No hay años válidos.")
        st.stop()

    años_sel = st.multiselect("Año", options=all_years, default=all_years)
    if not años_sel:
        st.warning("Selecciona al menos un año.")
        st.stop()

    # Filtrar por año
    df_ing = df_ing_raw[df_ing_raw["año"].isin(años_sel)].copy()
    df_egr = df_egr_raw[df_egr_raw["año"].isin(años_sel)].copy()

    # Rango de semanas (basado en ingresos + egresos)
    all_yw = sorted(set(
        df_ing["YEARWEEK"].dropna().astype(int).tolist() +
        df_egr["YEARWEEK"].dropna().astype(int).tolist()
    ))
    if len(all_yw) > 0:
        yw_labels = [yearweek_label(k) for k in all_yw]
        label_map = dict(zip(yw_labels, all_yw))

        if len(yw_labels) == 1:
            st.info(f"Semana: {yw_labels[0]}")
            yw_start, yw_end = all_yw[0], all_yw[0]
        else:
            lbl_start, lbl_end = st.select_slider(
                "Rango de semanas",
                options=yw_labels,
                value=(yw_labels[0], yw_labels[-1]),
            )
            yw_start, yw_end = label_map[lbl_start], label_map[lbl_end]
            if yw_start > yw_end:
                yw_start, yw_end = yw_end, yw_start

        df_ing = df_ing[
            (df_ing["YEARWEEK"].notna()) &
            (df_ing["YEARWEEK"] >= yw_start) &
            (df_ing["YEARWEEK"] <= yw_end)
        ]
        df_egr = df_egr[
            (df_egr["YEARWEEK"].notna()) &
            (df_egr["YEARWEEK"] >= yw_start) &
            (df_egr["YEARWEEK"] <= yw_end)
        ]

    # Filtro socio
    socios_all = sorted(set(
        [str(x) for x in df_ing["socio"].dropna().unique()] +
        [str(x) for x in df_egr["socio"].dropna().unique()]
    ))
    if socios_all:
        socio_sel = st.multiselect("Socio", options=socios_all, default=socios_all)
        if socio_sel:
            df_ing = df_ing[df_ing["socio"].isna() | df_ing["socio"].astype(str).isin(socio_sel)]
            df_egr = df_egr[df_egr["socio"].isna() | df_egr["socio"].astype(str).isin(socio_sel)]

    # Filtro conductor
    conductores_all = sorted(set(
        [str(x) for x in df_ing["conductor"].dropna().unique()]
    ))
    if conductores_all:
        search_cond = st.text_input("Buscar conductor")
        if search_cond.strip():
            patt = re.escape(search_cond.strip())
            conductores_f = [c for c in conductores_all if re.search(patt, c, re.IGNORECASE)]
        else:
            conductores_f = conductores_all
        cond_sel = st.multiselect("Conductor", options=conductores_f, default=conductores_f)
        if cond_sel:
            df_ing = df_ing[df_ing["conductor"].astype(str).isin(cond_sel)]


# ═══════════════════════════════════════════════════════════════
# SEPARACIÓN INVERSIÓN / GASTO OPERATIVO
# ──────────────────────────────────────────────────────────────
# df_gasto        → egresos sin INVERSIÓN ni PAGO DE SUBASTA
#                   (usar en KPIs de gasto, gráficas operativas, utilidad)
# df_inversion_egr→ solo egresos de tipo inversión
#                   (mostrar como KPI separado; NO resta la utilidad operativa)
# ═══════════════════════════════════════════════════════════════
df_gasto         = get_gastos(df_egr)
df_inversion_egr = get_inversiones_egreso(df_egr)


# ═══════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════
st.markdown("# 🚗 Dashboard Operación UBER")
st.caption(f"Datos cargados: **{len(df_ing):,}** registros de ingresos · **{len(df_egr):,}** registros de egresos")

# ═══════════════════════════════════════════════════════════════
# TABS
# ═══════════════════════════════════════════════════════════════
tab_ingresos, tab_egresos, tab_resumen, tab_vehiculos = st.tabs([
    "💰 Ingresos",
    "💸 Egresos",
    "📊 Resumen Global",
    "🚗 Vehículos",
])

# ═══════════════════════════════════════════════════════════════
# TAB 1: INGRESOS
# ═══════════════════════════════════════════════════════════════
with tab_ingresos:
    st.subheader("💰 Ingresos — Vista General")

    if len(df_ing) == 0:
        st.info("No hay datos de ingresos con los filtros actuales.")
    else:
        # ── KPIs con lógica de signos correcta (sin abs() ciego) ─────────
        total_renta       = df_ing["renta_semanal"].sum()
        total_fianza      = df_ing["fianza"].sum()
        total_ganancias   = df_ing["ganancias_totales"].sum()

        # Columnas neto (cargo - crédito): NO inflan por valores positivos
        total_multa_cargo    = df_ing["multa_cargo"].sum()       if "multa_cargo"    in df_ing.columns else 0
        total_multa_cred     = df_ing["multa_credito"].sum()     if "multa_credito"  in df_ing.columns else 0
        total_multa_neto     = df_ing["multa_neto"].sum()        if "multa_neto"     in df_ing.columns else 0

        total_hoj_cargo      = df_ing["hojalatero_cargo"].sum()   if "hojalatero_cargo"   in df_ing.columns else 0
        total_hoj_cred       = df_ing["hojalatero_credito"].sum() if "hojalatero_credito" in df_ing.columns else 0
        total_hoj_neto       = df_ing["hojalatero_neto"].sum()    if "hojalatero_neto"    in df_ing.columns else 0

        total_desc_cargo     = df_ing["descuentos_cargo"].sum()   if "descuentos_cargo"   in df_ing.columns else 0
        total_desc_cred      = df_ing["descuentos_credito"].sum() if "descuentos_credito" in df_ing.columns else 0
        total_desc_neto      = df_ing["descuentos_neto"].sum()    if "descuentos_neto"    in df_ing.columns else 0

        # Fila 1: KPIs principales
        k1, k2, k3 = st.columns(3)
        k1.metric("Renta Total",       f"${total_renta:,.0f}")
        k2.metric("Fianza Neta",       f"${total_fianza:,.0f}")
        k3.metric("Ganancias Totales", f"${total_ganancias:,.0f}")

        # Fila 2: KPIs neto (cargos - créditos)
        k4, k5, k6 = st.columns(3)
        k4.metric(
            "Multas Netas",
            f"${total_multa_neto:,.0f}",
            delta=f"Cargos ${total_multa_cargo:,.0f} · Créditos ${total_multa_cred:,.0f}",
            delta_color="off",
            help="Neto = Cargos cobrados − Créditos/devoluciones. Usa la columna multa_neto.",
        )
        k5.metric(
            "Hojalatero Neto",
            f"${total_hoj_neto:,.0f}",
            delta=f"Cargos ${total_hoj_cargo:,.0f} · Créditos ${total_hoj_cred:,.0f}",
            delta_color="off",
            help="Neto = Cargos − Créditos. Usa hojalatero_neto.",
        )
        k6.metric(
            "Descuentos Netos",
            f"${total_desc_neto:,.0f}",
            delta=f"Cargos ${total_desc_cargo:,.0f} · Créditos ${total_desc_cred:,.0f}",
            delta_color="off",
            help="Neto = Cargos − Créditos. El más crítico: muchos créditos positivos en DESC.",
        )

        st.divider()

        # Gráficas
        c1, c2 = st.columns(2)

        with c1:
            st.markdown("#### Renta Semanal por Semana")
            g = (
                df_ing.groupby("WEEK_LABEL", as_index=False)["renta_semanal"]
                .sum()
                .sort_values("WEEK_LABEL")
            )
            fig = px.bar(
                g, x="WEEK_LABEL", y="renta_semanal",
                color_discrete_sequence=[COLOR_INGRESOS],
            )
            fig.update_traces(hovertemplate="Semana: %{x}<br>Renta: $%{y:,.0f}<extra></extra>")
            fig.update_layout(xaxis_title="Semana", yaxis_title="Renta ($)")
            st.plotly_chart(styled_bar(fig), use_container_width=True)

        with c2:
            st.markdown("#### Ganancias Totales por Semana")
            g2 = (
                df_ing.groupby("WEEK_LABEL", as_index=False)["ganancias_totales"]
                .sum()
                .sort_values("WEEK_LABEL")
            )
            fig2 = px.line(
                g2, x="WEEK_LABEL", y="ganancias_totales",
                markers=True,
                color_discrete_sequence=[COLOR_NETO],
            )
            fig2.update_traces(hovertemplate="Semana: %{x}<br>Ganancias: $%{y:,.0f}<extra></extra>")
            fig2.update_layout(xaxis_title="Semana", yaxis_title="Ganancias ($)")
            st.plotly_chart(styled_line(fig2), use_container_width=True)

        # Renta por conductor (top 15)
        st.markdown("#### Top 15 Conductores — Renta Acumulada")
        top_cond = (
            df_ing.groupby("conductor", as_index=False)["renta_semanal"]
            .sum()
            .sort_values("renta_semanal", ascending=True)
            .tail(15)
        )
        fig3 = px.bar(
            top_cond, x="renta_semanal", y="conductor",
            orientation="h",
            color_discrete_sequence=[COLOR_INGRESOS],
        )
        fig3.update_traces(
            texttemplate="$%{x:,.0f}", textposition="outside",
            hovertemplate="%{y}<br>Renta: $%{x:,.0f}<extra></extra>",
        )
        fig3.update_layout(xaxis_title="Renta ($)", yaxis_title="")
        st.plotly_chart(styled_bar(fig3, money_axis="x"), use_container_width=True)

        # Distribución de conceptos
        c3, c4 = st.columns(2)
        with c3:
            st.markdown("#### Distribución de Conceptos de Ingreso")
            conc = df_ing["concepto_ingreso"].value_counts().reset_index()
            conc.columns = ["Concepto", "Registros"]
            fig4 = px.pie(
                conc.head(8), values="Registros", names="Concepto",
                color_discrete_sequence=COLOR_PALETTE,
                hole=0.4,
            )
            fig4.update_layout(
                template=PLOT_TEMPLATE,
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#c8d6e5"),
            )
            st.plotly_chart(fig4, use_container_width=True)

        with c4:
            st.markdown("#### Fianza Neta por Semana")
            g_fi = (
                df_ing.groupby("WEEK_LABEL", as_index=False)["fianza"]
                .sum()
                .sort_values("WEEK_LABEL")
            )
            fig5 = px.bar(
                g_fi, x="WEEK_LABEL", y="fianza",
                color_discrete_sequence=["#ffd93d"],
            )
            fig5.update_traces(hovertemplate="Semana: %{x}<br>Fianza: $%{y:,.0f}<extra></extra>")
            fig5.update_layout(xaxis_title="Semana", yaxis_title="Fianza Neta ($)")
            st.plotly_chart(styled_bar(fig5), use_container_width=True)

        # ── Conductores únicos por semana ─────────────────────────────────
        st.markdown("#### 👷 Conductores Únicos por Semana")
        st.caption(
            "Cada conductor se cuenta **una sola vez por semana**, "
            "eliminando duplicados del Excel antes de agregar."
        )
        cond_sem_df = conductores_por_semana(df_ing)
        if len(cond_sem_df) > 0:
            fig_cond = px.bar(
                cond_sem_df,
                x="WEEK_LABEL",
                y="n_conductores",
                color_discrete_sequence=["#a78bfa"],
                text="n_conductores",
            )
            fig_cond.update_traces(
                textposition="outside",
                hovertemplate="Semana: %{x}<br>Conductores: %{y}<extra></extra>",
            )
            fig_cond.update_layout(
                xaxis_title="Semana",
                yaxis_title="Conductores únicos",
                yaxis=dict(tickformat="d"),
            )
            st.plotly_chart(styled_bar(fig_cond, money_axis=None), use_container_width=True)

            # KPI rápido debajo de la gráfica
            prom_cond = cond_sem_df["n_conductores"].mean()
            max_cond  = cond_sem_df["n_conductores"].max()
            min_cond  = cond_sem_df["n_conductores"].min()
            kc1, kc2, kc3 = st.columns(3)
            kc1.metric("Prom. Conductores / Semana", f"{prom_cond:.1f}")
            kc2.metric("Máx. Conductores en una Semana", f"{int(max_cond)}")
            kc3.metric("Mín. Conductores en una Semana", f"{int(min_cond)}")
        else:
            st.info("No hay datos suficientes para graficar conductores por semana.")

        # Tabla detalle
        st.markdown("#### 📋 Tabla de Detalle — Ingresos")
        cols_show = [
            c for c in [
                "año", "semana", "WEEK_LABEL", "conductor", "llave", "socio",
                "app", "renta_semanal", "fianza",
                "multa_cargo", "multa_credito", "multa_neto",
                "hojalatero_cargo", "hojalatero_credito", "hojalatero_neto",
                "descuentos_cargo", "descuentos_credito", "descuentos_neto",
                "ganancias_totales", "concepto_ingreso",
            ] if c in df_ing.columns
        ]
        st.dataframe(
            df_ing[cols_show].sort_values(["año", "semana"], ascending=[True, True]),
            use_container_width=True,
            hide_index=True,
            height=400,
            column_config={
                "renta_semanal":      st.column_config.NumberColumn("Renta",      format="$%,.0f"),
                "fianza":             st.column_config.NumberColumn("Fianza",     format="$%,.0f"),
                "multa_cargo":        st.column_config.NumberColumn("Multa Cargo",   format="$%,.0f"),
                "multa_credito":      st.column_config.NumberColumn("Multa Créd.",   format="$%,.0f"),
                "multa_neto":         st.column_config.NumberColumn("Multa Neto",    format="$%,.0f"),
                "hojalatero_cargo":   st.column_config.NumberColumn("Hoj. Cargo",    format="$%,.0f"),
                "hojalatero_credito": st.column_config.NumberColumn("Hoj. Créd.",    format="$%,.0f"),
                "hojalatero_neto":    st.column_config.NumberColumn("Hoj. Neto",     format="$%,.0f"),
                "descuentos_cargo":   st.column_config.NumberColumn("Desc. Cargo",   format="$%,.0f"),
                "descuentos_credito": st.column_config.NumberColumn("Desc. Créd.",   format="$%,.0f"),
                "descuentos_neto":    st.column_config.NumberColumn("Desc. Neto",    format="$%,.0f"),
                "ganancias_totales":  st.column_config.NumberColumn("Gan. Totales",  format="$%,.0f"),
            },
        )

        # ── SECCIÓN: AUDITORÍA DE SIGNOS Y CONCEPTOS ─────────────────────
        st.divider()
        st.markdown("#### 🔍 Auditoría de Signos y Conceptos")
        st.caption(
            "Esta tabla permite validar cada KPI contra el Excel. "
            "**Suma Abs** es lo que mostraba el dashboard anterior (con `.abs()`). "
            "**Neto** es el valor correcto (cargos − créditos)."
        )

        aud_df = auditoria_signos(df_ing)
        if not aud_df.empty:
            st.dataframe(
                aud_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Concepto":        st.column_config.TextColumn("Concepto"),
                    "Suma Abs":        st.column_config.NumberColumn("Suma Abs (anterior)", format="$%,.2f"),
                    "Cargos (neg)":    st.column_config.NumberColumn("Cargos",   format="$%,.2f"),
                    "Creditos (pos)":  st.column_config.NumberColumn("Créditos", format="$%,.2f"),
                    "Neto":            st.column_config.NumberColumn("Neto ✓",   format="$%,.2f"),
                    "Filas negativas": st.column_config.NumberColumn("Filas neg"),
                    "Filas positivas": st.column_config.NumberColumn("Filas pos"),
                    "Filas nulas":     st.column_config.NumberColumn("Nulas"),
                },
            )
            # Resaltar diferencia más crítica (DESC)
            if len(aud_df) > 0:
                desc_row = aud_df[aud_df["Concepto"] == "Descuentos"]
                if not desc_row.empty:
                    inflacion = float(desc_row["Suma Abs"].iloc[0]) - float(desc_row["Neto"].iloc[0])
                    if inflacion > 0:
                        st.warning(
                            f"⚠️ **Descuentos**: el método anterior (abs) sobreestimaba en **${inflacion:,.2f}** "
                            f"({float(desc_row['Creditos (pos)'].iloc[0]):,.2f} de créditos contados como cargos)."
                        )


# ═══════════════════════════════════════════════════════════════
# TAB 2: EGRESOS
# ═══════════════════════════════════════════════════════════════
with tab_egresos:
    st.subheader("💸 Egresos — Vista General")

    if len(df_egr) == 0:
        st.info("No hay datos de egresos con los filtros actuales.")
    else:
        # ── KPIs de Gasto Operativo (excluye INVERSIÓN) ──────────────────
        # Métrica: monto_neto = gasto_bruto − créditos/ajustes
        # NO se usa .abs() ciegamente. Si un registro es negativo
        # (crédito/devolución), reduce el total correctamente.
        _col_neto  = "monto_neto"  if "monto_neto"  in df_gasto.columns else "monto_real"
        _col_bruto = "monto_gasto_bruto" if "monto_gasto_bruto" in df_gasto.columns else _col_neto
        _col_cred  = "monto_credito"     if "monto_credito"     in df_gasto.columns else None

        total_gasto_op   = df_gasto[_col_neto].sum()  if len(df_gasto) > 0 else 0.0
        total_inversion  = df_inversion_egr["monto_real"].sum() if len(df_inversion_egr) > 0 else 0.0
        n_semanas_egr    = df_gasto["YEARWEEK"].nunique() if len(df_gasto) > 0 else 0
        prom_semanal_op  = total_gasto_op / max(n_semanas_egr, 1)

        # Top de concepto — dinámico, usa la misma métrica que el total
        top_concepto = None
        if "concepto" in df_gasto.columns and len(df_gasto) > 0:
            top_concepto = (
                df_gasto.groupby("concepto", as_index=False)[_col_neto]
                .sum()
                .sort_values(_col_neto, ascending=False)
                .iloc[0]
            )

        st.markdown("##### 💡 Gasto Operativo *(INVERSIÓN excluida)*")
        st.caption(
            f"🔗 Métrica: `{_col_neto}` (gastos brutos − ajustes/créditos). "
            "El KPI de Top de Concepto usa la misma métrica que el Total y cambia dinámicamente con los filtros."
        )
        k1, k2, k3, k4 = st.columns(4)
        _help_total = (
            f"Gasto bruto: ${df_gasto[_col_bruto].sum():,.0f}\n"
            f"Créditos/ajustes: -${df_gasto[_col_cred].sum():,.0f}\n"
            f"Neto = bruto − créditos"
        ) if _col_cred else "Suma de monto_neto en gasto operativo"
        k1.metric("Gasto Operativo Total", f"${total_gasto_op:,.0f}", help=_help_total)
        k2.metric("Promedio Semanal", f"${prom_semanal_op:,.0f}",
                  help="Total gasto neto ÷ semanas distintas en el período filtrado")
        k3.metric("Semanas", f"{n_semanas_egr}")
        if top_concepto is not None:
            k4.metric(
                f"Top: {top_concepto['concepto']}",
                f"${top_concepto[_col_neto]:,.0f}",
                help="Concepto con mayor gasto neto. Cambia dinámicamente según el filtro activo.",
            )

        # KPI de inversión al lado, diferenciado
        st.markdown("##### 🏦 Inversión *(separada del gasto operativo)*")
        ki_col, ki_blank1, ki_blank2, ki_blank3 = st.columns(4)
        ki_col.metric(
            "💰 Inversión Total",
            f"${total_inversion:,.0f}",
            delta="No resta la utilidad operativa",
            delta_color="off",
        )

        st.divider()

        c1, c2 = st.columns(2)

        with c1:
            st.markdown("#### Gasto por Concepto — Operativo *(sin Inversión)*")
            if "concepto" in df_gasto.columns and len(df_gasto) > 0:
                by_conc = (
                    df_gasto.groupby("concepto", as_index=False)["monto_real"]
                    .sum()
                    .sort_values("monto_real", ascending=True)
                    .tail(15)
                )
                fig = px.bar(
                    by_conc, x="monto_real", y="concepto",
                    orientation="h",
                    color_discrete_sequence=[COLOR_EGRESOS],
                )
                fig.update_traces(
                    texttemplate="$%{x:,.0f}", textposition="outside",
                    hovertemplate="%{y}<br>Gasto: $%{x:,.0f}<extra></extra>",
                )
                fig.update_layout(xaxis_title="Gasto ($)", yaxis_title="")
                st.plotly_chart(styled_bar(fig, money_axis="x"), use_container_width=True)
            else:
                st.info("Sin datos de gasto operativo.")

        with c2:
            st.markdown("#### Gasto Semanal — Operativo *(sin Inversión)*")
            if len(df_gasto) > 0:
                g_egr = (
                    df_gasto.groupby("WEEK_LABEL", as_index=False)["monto_real"]
                    .sum()
                    .sort_values("WEEK_LABEL")
                )
                fig2 = px.line(
                    g_egr, x="WEEK_LABEL", y="monto_real",
                    markers=True,
                    color_discrete_sequence=[COLOR_EGRESOS],
                )
                fig2.update_traces(hovertemplate="Semana: %{x}<br>Gasto: $%{y:,.0f}<extra></extra>")
                fig2.update_layout(xaxis_title="Semana", yaxis_title="Gasto ($)")
                st.plotly_chart(styled_line(fig2), use_container_width=True)
            else:
                st.info("Sin datos de gasto semanal.")

        # Detalle por tipo (solo gasto operativo)
        if "detalle" in df_gasto.columns and len(df_gasto) > 0:
            st.markdown("#### Gasto por Detalle — Operativo")
            ver_todo_det = st.checkbox("Ver todos los detalles", value=False, key="det_all")
            by_det = (
                df_gasto.groupby("detalle", as_index=False)["monto_real"]
                .sum()
                .sort_values("monto_real", ascending=False)
            )
            by_det_show = by_det if ver_todo_det else by_det.head(5)
            fig3 = px.bar(
                by_det_show, x="detalle", y="monto_real",
                color_discrete_sequence=[COLOR_EGRESOS],
            )
            fig3.update_traces(
                texttemplate="$%{y:,.0f}", textposition="outside",
                hovertemplate="%{x}<br>Gasto: $%{y:,.0f}<extra></extra>",
            )
            fig3.update_layout(xaxis_title="Detalle", yaxis_title="Gasto ($)")
            st.plotly_chart(styled_bar(fig3), use_container_width=True)

        # Tabla detalle — Gasto Operativo
        st.markdown("#### 📋 Tabla de Detalle — Gasto Operativo *(sin Inversión)*")
        cols_egr = [
            c for c in [
                "año", "semana", "WEEK_LABEL", "concepto", "detalle",
                "conductor", "llave", "socio", "metodo_pago",
                "monto_real", "comercio",
            ] if c in df_gasto.columns
        ]
        if len(df_gasto) > 0:
            st.dataframe(
                df_gasto[cols_egr].sort_values(["año", "semana"], ascending=[True, True]),
                use_container_width=True,
                hide_index=True,
                height=400,
                column_config={
                    "monto_real": st.column_config.NumberColumn("Monto Real", format="$%,.0f"),
                },
            )

        # ── Sección Inversión ─────────────────────────────────────────────
        st.divider()
        st.markdown("#### 🏦 Movimientos de Inversión")
        st.caption(
            f"Conceptos clasificados como inversión: "
            f"**{', '.join(CONCEPTOS_INVERSION)}** — "
            f"estos montos aparecen aquí y **no se suman al gasto operativo**."
        )
        if len(df_inversion_egr) > 0:
            cols_inv = [
                c for c in [
                    "año", "semana", "WEEK_LABEL", "concepto", "detalle",
                    "conductor", "llave", "socio", "monto_real", "comercio",
                ] if c in df_inversion_egr.columns
            ]
            st.dataframe(
                df_inversion_egr[cols_inv].sort_values(["año", "semana"], ascending=[True, True]),
                use_container_width=True,
                hide_index=True,
                height=300,
                column_config={
                    "monto_real": st.column_config.NumberColumn("Monto Inversión", format="$%,.0f"),
                },
            )
        else:
            st.info(
                "No se encontraron movimientos de inversión en el período seleccionado. "
                f"Conceptos que se clasifican como inversión: {', '.join(CONCEPTOS_INVERSION)}"
            )

        # ── SECCIÓN: AUDITORÍA DE EGRESOS ──────────────────────────────────
        st.divider()
        st.markdown("#### 🔍 Auditoría de Egresos")
        st.caption(
            "Esta tabla muestra el desglose por concepto: gastos positivos vs. ajustes/créditos negativos. "
            "**Suma absoluta** és lo que mostraba el dashboard anterior (con `.abs()`). "
            "**Neto** es el valor correcto (gastos brutos − ajustes)."
        )

        # Pasar el df de gasto ya filtrado (respeta los filtros activos)
        aud_egr_conc, aud_egr_sosp = auditoria_egresos(df_gasto)

        if not aud_egr_conc.empty:
            # Totales del pie
            total_row = pd.DataFrame([{
                "Concepto":         "⟹ TOTAL",
                "Suma raw firmada":  aud_egr_conc["Suma raw firmada"].sum(),
                "Suma absoluta":     aud_egr_conc["Suma absoluta"].sum(),
                "Gastos positivos":  aud_egr_conc["Gastos positivos"].sum(),
                "Ajustes/Créditos": aud_egr_conc["Ajustes/Créditos"].sum(),
                "Neto":              aud_egr_conc["Neto"].sum(),
                "Filas positivas":   aud_egr_conc["Filas positivas"].sum(),
                "Filas negativas":   aud_egr_conc["Filas negativas"].sum(),
                "Filas cero/nulas":  aud_egr_conc["Filas cero/nulas"].sum(),
            }])
            aud_egr_display = pd.concat([aud_egr_conc, total_row], ignore_index=True)

            st.dataframe(
                aud_egr_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Concepto":          st.column_config.TextColumn("Concepto"),
                    "Suma raw firmada":   st.column_config.NumberColumn("Suma Raw firmada",  format="$%,.2f"),
                    "Suma absoluta":      st.column_config.NumberColumn("Suma Abs (antes)",   format="$%,.2f"),
                    "Gastos positivos":   st.column_config.NumberColumn("Gastos positivos",   format="$%,.2f"),
                    "Ajustes/Créditos":  st.column_config.NumberColumn("Ajustes/Créditos",  format="$%,.2f"),
                    "Neto":               st.column_config.NumberColumn("Neto ✓",             format="$%,.2f"),
                    "Filas positivas":    st.column_config.NumberColumn("Filas pos"),
                    "Filas negativas":    st.column_config.NumberColumn("Filas neg"),
                    "Filas cero/nulas":   st.column_config.NumberColumn("Nulas"),
                },
            )

            # Resaltar inflación si existe
            total_abs  = float(aud_egr_conc["Suma absoluta"].sum())
            total_neto = float(aud_egr_conc["Neto"].sum())
            inflacion  = total_abs - total_neto
            total_cred = float(aud_egr_conc["Ajustes/Créditos"].sum())
            if inflacion > 0.01:
                st.warning(
                    f"⚠️ El método anterior (abs) sobreestimaba el gasto operativo en "
                    f"**${inflacion:,.2f}** debido a {int(aud_egr_conc['Filas negativas'].sum())} "
                    f"ajuste(s)/crédito(s) por **${total_cred:,.2f}** que se contaban como gastos."
                )
            else:
                st.success("✅ Sin diferencia entre suma absoluta y neto — no hay ajustes negativos en el período.")

        # Tabla de filas sospechosas
        if not aud_egr_sosp.empty:
            st.markdown("##### 🚨 Filas con monto negativo (ajustes/créditos)")
            st.caption(
                "Estas filas tienen un monto negativo en la columna REAL del Excel. "
                "Se interpretan como ajustes, devoluciones o créditos que **reducen** el gasto operativo."
            )
            st.dataframe(
                aud_egr_sosp,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "monto_original":    st.column_config.NumberColumn("Monto original",    format="$%,.2f"),
                    "monto_interpretado": st.column_config.NumberColumn("Neto interpretado", format="$%,.2f"),
                },
            )
        else:
            st.info("✅ No hay filas con monto negativo en el gasto operativo del período seleccionado.")
# ═══════════════════════════════════════════════════════════════
# TAB 3: RESUMEN GLOBAL
# ═══════════════════════════════════════════════════════════════
with tab_resumen:
    st.subheader("📊 Resumen Global — Ingresos vs Egresos")

    total_ing        = df_ing["ganancias_totales"].sum() if len(df_ing) > 0 else 0.0
    # Utilidad operativa: solo gasto operativo (SIN INVERSIÓN)
    # Usa monto_neto para consistencia con el tab de Egresos (no .abs() ciego)
    _col_neto_res    = "monto_neto" if "monto_neto" in df_gasto.columns else "monto_real"
    total_egr_op     = df_gasto[_col_neto_res].sum() if len(df_gasto) > 0 else 0.0
    total_inv_global = df_inversion_egr["monto_real"].sum() if len(df_inversion_egr) > 0 else 0.0
    utilidad         = total_ing - total_egr_op   # operativa, sin descontar inversión

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Ingresos Totales", f"${total_ing:,.0f}")
    k2.metric("Gasto Operativo", f"${total_egr_op:,.0f}", help="Excluye INVERSIÓN y PAGO DE SUBASTA")
    k3.metric(
        "Utilidad Operativa",
        f"${utilidad:,.0f}",
        delta=f"{'✅' if utilidad >= 0 else '⚠️'} {'Positiva' if utilidad >= 0 else 'Negativa'}",
    )
    k4.metric(
        "🏦 Inversión",
        f"${total_inv_global:,.0f}",
        delta="No resta la utilidad",
        delta_color="off",
        help="INVERSIÓN + PAGO DE SUBASTA — separados del gasto operativo",
    )

    st.divider()

    c1, c2 = st.columns(2)

    with c1:
        st.markdown("#### Ingreso vs Gasto Operativo por Semana")
        st.caption("Los Egresos aquí representan **solo gasto operativo** (sin Inversión)")

        # Agrupar ingresos por semana
        if len(df_ing) > 0:
            gi = (
                df_ing.groupby(["YEARWEEK", "WEEK_LABEL"], as_index=False)["ganancias_totales"]
                .sum()
                .rename(columns={"ganancias_totales": "Ingresos"})
            )
        else:
            gi = pd.DataFrame(columns=["YEARWEEK", "WEEK_LABEL", "Ingresos"])

        # Agrupar gasto operativo por semana (sin inversión)
        if len(df_gasto) > 0:
            ge = (
                df_gasto.groupby(["YEARWEEK", "WEEK_LABEL"], as_index=False)["monto_real"]
                .sum()
                .rename(columns={"monto_real": "Egresos"})
            )
        else:
            ge = pd.DataFrame(columns=["YEARWEEK", "WEEK_LABEL", "Egresos"])

        # Merge
        merged = pd.merge(gi, ge, on=["YEARWEEK", "WEEK_LABEL"], how="outer").fillna(0)
        merged = merged.sort_values("YEARWEEK")

        if len(merged) > 0:
            m = merged.melt(
                id_vars="WEEK_LABEL",
                value_vars=["Ingresos", "Egresos"],
                var_name="Tipo", value_name="Monto",
            )
            fig = px.bar(
                m, x="WEEK_LABEL", y="Monto", color="Tipo",
                barmode="group",
                color_discrete_map={"Ingresos": COLOR_INGRESOS, "Egresos": COLOR_EGRESOS},
            )
            fig.update_traces(hovertemplate="%{x}<br>%{data.name}: $%{y:,.0f}<extra></extra>")
            fig.update_layout(xaxis_title="Semana", yaxis_title="Monto ($)")
            st.plotly_chart(styled_bar(fig), use_container_width=True)

    with c2:
        st.markdown("#### Distribución de Gasto Operativo por Concepto")
        st.caption("Solo gasto operativo — Inversión excluida")
        if len(df_gasto) > 0 and "concepto" in df_gasto.columns:
            conc_egr = (
                df_gasto.groupby("concepto", as_index=False)["monto_real"]
                .sum()
                .sort_values("monto_real", ascending=False)
            )
            fig2 = px.pie(
                conc_egr.head(10), values="monto_real", names="concepto",
                color_discrete_sequence=COLOR_PALETTE,
                hole=0.45,
            )
            fig2.update_layout(
                template=PLOT_TEMPLATE,
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#c8d6e5"),
            )
            st.plotly_chart(fig2, use_container_width=True)

    # Utilidad por semana
    st.markdown("#### Utilidad Operativa por Semana *(Ingreso − Gasto Operativo)*")
    if len(merged) > 0:
        merged["Utilidad"] = merged["Ingresos"] - merged["Egresos"]
        colors = [COLOR_INGRESOS if v >= 0 else COLOR_EGRESOS for v in merged["Utilidad"]]

        fig3 = go.Figure(go.Bar(
            x=merged["WEEK_LABEL"],
            y=merged["Utilidad"],
            marker_color=colors,
            text=[f"${v:,.0f}" for v in merged["Utilidad"]],
            textposition="outside",
        ))
        fig3.update_layout(
            xaxis_title="Semana", yaxis_title="Utilidad ($)",
        )
        st.plotly_chart(styled_bar(fig3), use_container_width=True)

    # Tabla resumen por semana
    st.markdown("#### 📋 Resumen por Semana *(Gasto = solo operativo)*")
    if len(merged) > 0:
        merged_show = merged[["WEEK_LABEL", "Ingresos", "Egresos", "Utilidad"]].copy()
        merged_show = merged_show.sort_values("WEEK_LABEL")
        st.dataframe(
            merged_show,
            use_container_width=True,
            hide_index=True,
            height=350,
            column_config={
                "WEEK_LABEL": st.column_config.TextColumn("Semana"),
                "Ingresos": st.column_config.NumberColumn(format="$%,.0f"),
                "Egresos": st.column_config.NumberColumn("Gasto Operativo", format="$%,.0f"),
                "Utilidad": st.column_config.NumberColumn(format="$%,.0f"),
            },
        )


# ═══════════════════════════════════════════════════════════════
# TAB 4: VEHÍCULOS
# ═══════════════════════════════════════════════════════════════
with tab_vehiculos:
    st.subheader("🚗 Análisis por Vehículo")

    if len(df_ing) == 0:
        st.info("No hay datos de ingresos para analizar vehículos.")
    else:
        # ── KPIs GLOBALES DE FLOTA ──────────────────────────────────
        llaves_all = [str(x) for x in df_ing["llave"].dropna().unique() if str(x) != "-"]
        n_vehiculos = len(llaves_all)

        # Agrupar ingresos por vehículo (toda la flota)
        fleet_ing = (
            df_ing[df_ing["llave"].astype(str).isin(llaves_all)]
            .groupby("llave", as_index=False)
            .agg(
                Renta=("renta_semanal", "sum"),
                Ganancias=("ganancias_totales", "sum"),
                Semanas=("semana", "nunique"),
            )
        )

        # Agrupar gasto operativo por vehículo (SIN inversión)
        fleet_egr = pd.DataFrame(columns=["llave", "Egresos"])
        if len(df_gasto) > 0 and "llave" in df_gasto.columns:
            fleet_egr = (
                df_gasto[df_gasto["llave"].astype(str).isin(llaves_all)]
                .groupby("llave", as_index=False)["monto_real"]
                .sum()
                .rename(columns={"monto_real": "Egresos"})
            )

        fleet = fleet_ing.merge(fleet_egr, on="llave", how="left").fillna(0)
        fleet["Utilidad"] = fleet["Ganancias"] - fleet["Egresos"]

        total_fleet_ing = fleet["Ganancias"].sum()
        total_fleet_egr = fleet["Egresos"].sum()
        total_fleet_util = fleet["Utilidad"].sum()
        avg_ing_veh = total_fleet_ing / max(n_vehiculos, 1)
        avg_egr_veh = total_fleet_egr / max(n_vehiculos, 1)
        avg_semanas = fleet["Semanas"].mean() if len(fleet) > 0 else 0
        best_veh = fleet.sort_values("Utilidad", ascending=False).iloc[0] if len(fleet) > 0 else None
        worst_veh = fleet.sort_values("Utilidad", ascending=True).iloc[0] if len(fleet) > 0 else None
        veh_positivos = int((fleet["Utilidad"] > 0).sum())
        margen_pct = (total_fleet_util / total_fleet_ing * 100) if total_fleet_ing > 0 else 0

        st.markdown("#### 📊 Panorama General de Flota")

        # Fila 1: Métricas principales
        f1, f2, f3, f4 = st.columns(4)
        f1.metric("🚗 Vehículos Activos", f"{n_vehiculos}")
        f2.metric("💰 Ingresos Flota", f"${total_fleet_ing:,.0f}")
        f3.metric("💸 Gasto Op. Flota", f"${total_fleet_egr:,.0f}", help="Sin Inversión")
        f4.metric(
            "📈 Utilidad Operativa Flota",
            f"${total_fleet_util:,.0f}",
            delta=f"{'✅' if total_fleet_util >= 0 else '⚠️'} Margen {margen_pct:.1f}%",
        )

        # Fila 2: Promedios y top/bottom
        f5, f6, f7, f8 = st.columns(4)
        f5.metric("Ingreso Prom / Vehículo", f"${avg_ing_veh:,.0f}")
        f6.metric("Gasto Prom / Vehículo", f"${avg_egr_veh:,.0f}")
        f7.metric("📅 Semanas Prom Activo", f"{avg_semanas:.1f}")
        f8.metric("✅ Vehículos Rentables", f"{veh_positivos} / {n_vehiculos}")

        # Fila 3: Mejor y peor vehículo
        if best_veh is not None and worst_veh is not None:
            b1, b2 = st.columns(2)
            b1.metric(
                f"🏆 Mejor Vehículo: {best_veh['llave']}",
                f"${best_veh['Utilidad']:,.0f} utilidad",
                delta=f"${best_veh['Ganancias']:,.0f} ingresos / ${best_veh['Egresos']:,.0f} egresos",
            )
            b2.metric(
                f"⚠️ Menor Rendimiento: {worst_veh['llave']}",
                f"${worst_veh['Utilidad']:,.0f} utilidad",
                delta=f"${worst_veh['Ganancias']:,.0f} ingresos / ${worst_veh['Egresos']:,.0f} egresos",
                delta_color="inverse",
            )

        # Top 10 vehículos por utilidad (gráfica rápida)
        st.markdown("#### 🏅 Top 10 Vehículos por Utilidad")
        top10 = fleet.sort_values("Utilidad", ascending=True).tail(10)
        colors_top = [COLOR_INGRESOS if v >= 0 else COLOR_EGRESOS for v in top10["Utilidad"]]
        fig_top = go.Figure(go.Bar(
            x=top10["Utilidad"],
            y=top10["llave"],
            orientation="h",
            marker_color=colors_top,
            text=[f"${v:,.0f}" for v in top10["Utilidad"]],
            textposition="outside",
        ))
        fig_top.update_layout(xaxis_title="Utilidad ($)", yaxis_title="")
        st.plotly_chart(styled_bar(fig_top, money_axis="x"), use_container_width=True)

        st.divider()
        st.markdown("#### 🔍 Análisis Individual por Vehículo")

        # Selector de llaves
        llaves = sorted(llaves_all)

        search_ll = st.text_input("🔍 Buscar vehículo (llave)", key="veh_search")
        if search_ll.strip():
            patt = re.escape(search_ll.strip())
            llaves_f = [l for l in llaves if re.search(patt, l, re.IGNORECASE)]
        else:
            llaves_f = llaves

        llave_sel = st.multiselect(
            "Selecciona vehículo(s)",
            options=llaves_f,
            default=llaves_f[:3] if len(llaves_f) > 0 else [],
            key="veh_llave",
        )

        if not llave_sel:
            st.info("Selecciona al menos un vehículo para ver el análisis.")
        else:
            df_v = df_ing[df_ing["llave"].astype(str).isin(llave_sel)].copy()

            # KPI por vehículo — datos
            kpi_v = (
                df_v.groupby("llave", as_index=False)
                .agg(
                    Renta=("renta_semanal", "sum"),
                    Fianza=("fianza", "sum"),
                    Multas=("multa_neto", "sum"),
                    Hojalatero=("hojalatero_neto", "sum"),
                    Descuentos=("descuentos_neto", "sum"),
                    Ganancias=("ganancias_totales", "sum"),
                    Semanas=("semana", "nunique"),
                )
                .sort_values("Ganancias", ascending=False)
            )

            # Agregar gasto operativo por vehículo (SIN inversión)
            if len(df_gasto) > 0 and "llave" in df_gasto.columns:
                egr_v = (
                    df_gasto[df_gasto["llave"].astype(str).isin(llave_sel)]
                    .groupby("llave", as_index=False)["monto_real"]
                    .sum()
                    .rename(columns={"monto_real": "Egresos"})
                )
                kpi_v = kpi_v.merge(egr_v, on="llave", how="left").fillna(0)
                kpi_v["Utilidad"] = kpi_v["Ganancias"] - kpi_v["Egresos"]

            # ── KPI CARDS para vehículo(s) seleccionado(s) ──────────
            sel_renta       = kpi_v["Renta"].sum()
            sel_fianza      = kpi_v["Fianza"].sum()
            sel_multas      = kpi_v["Multas"].sum()
            sel_hojalatero  = kpi_v["Hojalatero"].sum()
            sel_descuentos  = kpi_v["Descuentos"].sum() if "Descuentos" in kpi_v.columns else 0
            sel_ganancias   = kpi_v["Ganancias"].sum()
            sel_semanas     = kpi_v["Semanas"].sum()
            sel_egresos     = kpi_v["Egresos"].sum() if "Egresos" in kpi_v.columns else 0
            sel_utilidad    = kpi_v["Utilidad"].sum() if "Utilidad" in kpi_v.columns else sel_ganancias

            st.markdown(f"#### 📊 KPIs — {', '.join(llave_sel[:3])}{'…' if len(llave_sel) > 3 else ''}")

            v1, v2, v3, v4 = st.columns(4)
            v1.metric("💰 Renta Total",      f"${sel_renta:,.0f}")
            v2.metric("📈 Ganancias Totales", f"${sel_ganancias:,.0f}")
            v3.metric("💸 Gasto Operativo",   f"${sel_egresos:,.0f}", help="Sin Inversión")
            v4.metric(
                "📊 Utilidad Operativa",
                f"${sel_utilidad:,.0f}",
                delta=f"{'✅ Positiva' if sel_utilidad >= 0 else '⚠️ Negativa'}",
            )

            v5, v6, v7, v8 = st.columns(4)
            v5.metric("🔒 Fianza Neta",    f"${sel_fianza:,.0f}")
            v6.metric("🚨 Multas Netas",   f"${sel_multas:,.0f}",     help="Cargos − Créditos")
            v7.metric("🔧 Hojalatero Neto",f"${sel_hojalatero:,.0f}", help="Cargos − Créditos")
            v8.metric("🏷️ Desc. Netos",    f"${sel_descuentos:,.0f}", help="Cargos − Créditos")

            st.divider()

            # Tabla rendimiento por vehículo
            st.markdown("#### Rendimiento por Vehículo")
            st.dataframe(
                kpi_v,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "llave": st.column_config.TextColumn("Vehículo"),
                    "Renta": st.column_config.NumberColumn(format="$%,.0f"),
                    "Fianza": st.column_config.NumberColumn(format="$%,.0f"),
                    "Multas": st.column_config.NumberColumn(format="$%,.0f"),
                    "Hojalatero": st.column_config.NumberColumn(format="$%,.0f"),
                    "Ganancias": st.column_config.NumberColumn(format="$%,.0f"),
                    "Egresos": st.column_config.NumberColumn(format="$%,.0f"),
                    "Utilidad": st.column_config.NumberColumn(format="$%,.0f"),
                },
            )

            st.divider()

            # Gráfica temporal por vehículo
            st.markdown("#### Rendimiento Semanal por Vehículo")
            metric_opt = st.radio(
                "Métrica:",
                ["Ganancias Totales", "Renta Semanal", "Fianza"],
                horizontal=True,
            )
            ycol_map = {
                "Ganancias Totales": "ganancias_totales",
                "Renta Semanal": "renta_semanal",
                "Fianza": "fianza",
            }
            ycol = ycol_map[metric_opt]

            g_v = (
                df_v.groupby(["YEARWEEK", "WEEK_LABEL", "llave"], as_index=False)[ycol]
                .sum()
                .sort_values(["YEARWEEK", "llave"])
            )
            fig = px.line(
                g_v, x="WEEK_LABEL", y=ycol, color="llave",
                markers=True,
                color_discrete_sequence=COLOR_PALETTE,
            )
            fig.update_traces(hovertemplate="Semana: %{x}<br>$%{y:,.0f}<extra></extra>")
            fig.update_layout(xaxis_title="Semana", yaxis_title=metric_opt)
            st.plotly_chart(styled_line(fig), use_container_width=True)

            # Historial de egresos por vehículo
            if len(df_egr) > 0 and "llave" in df_egr.columns:
                df_egr_v = df_egr[df_egr["llave"].astype(str).isin(llave_sel)].copy()
                if len(df_egr_v) > 0:
                    st.markdown("#### Historial de Egresos del Vehículo")
                    cols_ev = [
                        c for c in [
                            "año", "semana", "WEEK_LABEL", "concepto", "detalle",
                            "llave", "conductor", "monto_real", "comercio",
                        ] if c in df_egr_v.columns
                    ]
                    st.dataframe(
                        df_egr_v[cols_ev].sort_values(["año", "semana"]),
                        use_container_width=True,
                        hide_index=True,
                        height=350,
                        column_config={
                            "monto_real": st.column_config.NumberColumn("Monto", format="$%,.0f"),
                        },
                    )

            # Detalle ingresos
            st.markdown("#### 📋 Detalle de Ingresos por Vehículo")
            cols_v = [
                c for c in [
                    "año", "semana", "WEEK_LABEL", "conductor", "llave",
                    "renta_semanal", "fianza",
                    "multa_cargo", "multa_credito", "multa_neto",
                    "hojalatero_cargo", "hojalatero_credito", "hojalatero_neto",
                    "descuentos_cargo", "descuentos_credito", "descuentos_neto",
                    "ganancias_totales", "concepto_ingreso",
                ] if c in df_v.columns
            ]
            st.dataframe(
                df_v[cols_v].sort_values(["año", "semana"]),
                use_container_width=True,
                hide_index=True,
                height=400,
                column_config={
                    "renta_semanal":      st.column_config.NumberColumn("Renta",        format="$%,.0f"),
                    "fianza":             st.column_config.NumberColumn("Fianza",       format="$%,.0f"),
                    "multa_cargo":        st.column_config.NumberColumn("Multa Cargo",  format="$%,.0f"),
                    "multa_credito":      st.column_config.NumberColumn("Multa Créd.",  format="$%,.0f"),
                    "multa_neto":         st.column_config.NumberColumn("Multa Neto",   format="$%,.0f"),
                    "hojalatero_cargo":   st.column_config.NumberColumn("Hoj. Cargo",   format="$%,.0f"),
                    "hojalatero_credito": st.column_config.NumberColumn("Hoj. Créd.",   format="$%,.0f"),
                    "hojalatero_neto":    st.column_config.NumberColumn("Hoj. Neto",    format="$%,.0f"),
                    "descuentos_cargo":   st.column_config.NumberColumn("Desc. Cargo",  format="$%,.0f"),
                    "descuentos_credito": st.column_config.NumberColumn("Desc. Créd.",  format="$%,.0f"),
                    "descuentos_neto":    st.column_config.NumberColumn("Desc. Neto",   format="$%,.0f"),
                    "ganancias_totales":  st.column_config.NumberColumn("Gan. Totales", format="$%,.0f"),
                },
            )

        # ════════════════════════════════════════════════════════
        # ANÁLISIS DE INVERSIÓN — RECUPERACIÓN DE SUBASTA
        # ════════════════════════════════════════════════════════
        st.divider()
        st.markdown("### 💰 Análisis de Inversión — Recuperación de Subasta")
        st.caption(
            "Modela la recuperación del capital invertido en la compra de vehículos "
            "(concepto **PAGO DE SUBASTA**) mediante pagos semanales fijos de "
            f"**${PAGO_SEMANAL_DEFAULT:,.0f} MXN**."
        )

        # ── Cargar análisis (basado en egresos SIN filtro de llave para tener todo el historial)
        # Usamos df_egr_raw filtrado por año pero NO por llave para conservar subastas históricas
        df_egr_para_inv = df_egr_raw[df_egr_raw["año"].isin(años_sel)].copy()
        df_ing_para_inv = df_ing_raw[df_ing_raw["año"].isin(años_sel)].copy()

        inv_data = get_analisis_inversiones(df_egr_para_inv, df_ing_para_inv)
        resumen_inv = inv_data["resumen"]
        tablas_inv  = inv_data["tablas"]

        if resumen_inv.empty:
            st.info(
                "No se encontraron registros con concepto **PAGO DE SUBASTA** "
                "en los egresos del período seleccionado."
            )
        else:
            # ── KPIs de flota — inversiones ─────────────────────────────
            total_capital    = resumen_inv["capital_inicial"].sum()
            prom_semanas_rec = resumen_inv["semanas_recuperacion"].mean()
            n_recuperados    = int(resumen_inv["recuperado"].sum())
            n_inv            = len(resumen_inv)
            roi_prom         = resumen_inv["roi_pct"].mean()

            ki1, ki2, ki3, ki4 = st.columns(4)
            ki1.metric("💵 Capital Total Invertido",  f"${total_capital:,.0f}")
            ki2.metric("📅 Semanas Prom. Recuperación", f"{prom_semanas_rec:.1f}")
            ki3.metric("✅ Vehículos Recuperados",     f"{n_recuperados} / {n_inv}")
            ki4.metric(
                "📈 ROI Promedio Flota",
                f"{roi_prom:.1f}%",
                delta="Positivo" if roi_prom >= 0 else "Negativo",
                delta_color="normal" if roi_prom >= 0 else "inverse",
            )

            st.divider()

            # ── Gráfica comparativa ROI ─────────────────────────────────
            st.markdown("#### 📊 ROI Estimado por Vehículo")
            roi_df = resumen_inv.sort_values("roi_pct", ascending=True)
            colors_roi = [COLOR_INGRESOS if v >= 0 else COLOR_EGRESOS for v in roi_df["roi_pct"]]
            fig_roi = go.Figure(go.Bar(
                x=roi_df["roi_pct"],
                y=roi_df["llave"],
                orientation="h",
                marker_color=colors_roi,
                text=[f"{v:.1f}%" for v in roi_df["roi_pct"]],
                textposition="outside",
                hovertemplate="%{y}<br>ROI: %{x:.1f}%<extra></extra>",
            ))
            fig_roi.update_layout(xaxis_title="ROI (%)", yaxis_title="")
            st.plotly_chart(styled_bar(fig_roi, money_axis=None), use_container_width=True)

            st.divider()

            # ── Selector individual ─────────────────────────────────────
            st.markdown("#### 🔍 Detalle por Vehículo")
            llaves_inv = resumen_inv["llave"].tolist()
            llave_inv_sel = st.selectbox(
                "Selecciona el vehículo para ver su tabla de amortización:",
                options=llaves_inv,
                key="inv_llave_sel",
            )

            if llave_inv_sel:
                row = resumen_inv[resumen_inv["llave"] == llave_inv_sel].iloc[0]
                tabla_am = tablas_inv.get(llave_inv_sel, None)

                # Métricas individuales
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("💵 Capital Inicial",       f"${row['capital_inicial']:,.0f}")
                m2.metric("📆 Semanas p/Recuperar",   f"{row['semanas_recuperacion']}")
                m3.metric("💰 Ingresos Totales",       f"${row['ingresos_totales']:,.0f}")
                m4.metric(
                    "📈 ROI Estimado",
                    f"{row['roi_pct']:.1f}%",
                    delta="✅ Inversión recuperada" if row["recuperado"] else "⏳ En recuperación",
                    delta_color="normal" if row["recuperado"] else "off",
                )

                if tabla_am is not None and not tabla_am.empty:
                    c_tabla, c_graf = st.columns([1, 1])

                    with c_tabla:
                        st.markdown("##### 📋 Tabla de Amortización")
                        tabla_show = tabla_am.rename(columns={
                            "semana_num":        "Semana",
                            "pago_recuperacion":  "Pago Recuperación",
                            "saldo_restante":     "Saldo Restante",
                        })
                        st.dataframe(
                            tabla_show,
                            use_container_width=True,
                            hide_index=True,
                            height=320,
                            column_config={
                                "Pago Recuperación": st.column_config.NumberColumn(format="$%,.0f"),
                                "Saldo Restante":    st.column_config.NumberColumn(format="$%,.0f"),
                            },
                        )

                    with c_graf:
                        st.markdown("##### 📉 Curva de Recuperación del Capital")
                        # Semana de break-even (saldo llega a 0)
                        breakeven_sem = int(tabla_am[tabla_am["saldo_restante"] == 0]["semana_num"].min())

                        fig_am = go.Figure()

                        # Área bajo la curva
                        fig_am.add_trace(go.Scatter(
                            x=tabla_am["semana_num"],
                            y=tabla_am["saldo_restante"],
                            mode="lines",
                            name="Saldo Restante",
                            line=dict(color="#ff6b6b", width=2.5),
                            fill="tozeroy",
                            fillcolor="rgba(255,107,107,0.15)",
                            hovertemplate="Semana %{x}<br>Saldo: $%{y:,.0f}<extra></extra>",
                        ))

                        # Línea vertical break-even
                        fig_am.add_vline(
                            x=breakeven_sem,
                            line_dash="dot",
                            line_color="#00d4aa",
                            annotation_text=f"Break-even S{breakeven_sem}",
                            annotation_position="top right",
                            annotation_font_color="#00d4aa",
                        )

                        # Línea horizontal en 0
                        fig_am.add_hline(
                            y=0,
                            line_dash="dash",
                            line_color="#a8b8d8",
                            line_width=1,
                        )

                        fig_am.update_layout(
                            xaxis_title="Semana de pago",
                            yaxis_title="Saldo Pendiente ($)",
                            yaxis_tickformat="$,.0f",
                        )
                        st.plotly_chart(styled_line(fig_am, money_axis="y"), use_container_width=True)

            # ── Tabla resumen de todas las inversiones ──────────────────
            st.markdown("#### 📋 Resumen de Inversiones en Flota")
            resumen_show = resumen_inv.rename(columns={
                "llave":                 "Vehículo",
                "capital_inicial":       "Capital Inicial",
                "pago_semanal":          "Pago Semanal",
                "semanas_recuperacion":  "Semanas p/Recuperar",
                "monto_recuperado":      "Monto Recuperado",
                "ingresos_totales":      "Ingresos Totales",
                "semanas_activas":       "Semanas Activas",
                "roi_pct":               "ROI (%)",
                "recuperado":            "¿Recuperado?",
            })
            st.dataframe(
                resumen_show,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Capital Inicial":      st.column_config.NumberColumn(format="$%,.0f"),
                    "Pago Semanal":         st.column_config.NumberColumn(format="$%,.0f"),
                    "Monto Recuperado":     st.column_config.NumberColumn(format="$%,.0f"),
                    "Ingresos Totales":     st.column_config.NumberColumn(format="$%,.0f"),
                    "ROI (%)": st.column_config.NumberColumn(format="%.1f%%"),
                    "¿Recuperado?": st.column_config.CheckboxColumn(),
                },
            )

            # ── Botón de descarga Excel ───────────────────────────
            st.divider()
            st.markdown("#### 📥 Exportar a Excel")

            # Armar dict de KPIs globales para la hoja KPIs
            saldo_pendiente_total = float(
                resumen_inv["capital_inicial"].sum()
                - resumen_inv["monto_recuperado"].sum()
            )
            total_recuperado = float(resumen_inv["monto_recuperado"].sum())

            kpis_export = {
                "Capital Total Invertido":       {"valor": float(total_capital),          "tipo": "moneda"},
                "Total Recuperado":              {"valor": total_recuperado,               "tipo": "moneda"},
                "Saldo Pendiente Total":         {"valor": saldo_pendiente_total,          "tipo": "moneda"},
                "Semanas Prom. Recuperación":   {"valor": float(prom_semanas_rec),        "tipo": "numero"},
                "Vehículos Analizados":          {"valor": int(n_inv),                     "tipo": "numero"},
                "Vehículos Recuperados":         {"valor": int(n_recuperados),             "tipo": "numero"},
                "ROI Promedio Flota (%)": {"valor": float(roi_prom),               "tipo": "porcentaje"},
                "Pago Semanal Fijo":             {"valor": float(PAGO_SEMANAL_DEFAULT),   "tipo": "moneda"},
            }

            try:
                excel_bytes = _generar_excel_amortizacion(
                    resumen_inv=resumen_inv,
                    tablas_inv=tablas_inv,
                    df_egr_para_inv=df_egr_para_inv,
                    kpis=kpis_export,
                )
                st.download_button(
                    label="📥 Descargar Análisis de Inversión (.xlsx)",
                    data=excel_bytes,
                    file_name="analisis_inversion_vehiculos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as exc:
                st.error(f"No se pudo generar el Excel: {exc}")
